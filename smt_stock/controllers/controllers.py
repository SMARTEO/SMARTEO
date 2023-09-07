#-*- coding: utf-8 -*-
from collections import deque
import io
import json
import base64
from odoo import http, _
from odoo.http import content_disposition, request
from odoo.tools import ustr, osutil
from odoo.tools.misc import xlsxwriter
from openpyxl import Workbook
from werkzeug.wrappers import Response
from openpyxl.styles import Font

import functools

SUB_FIELDS_TO_GET = {
    'sale_id': 'name',
    'move_lines': ['name', 'product_qty']
}

class SmtStock(http.Controller):

    # @http.route('/export', type='http', auth='user')
    # def export_handler(self, **kw):
    #     model = request.env['stock.picking']
    #     response = model.export_data_planning_issu()
    #     return response(environ=request.httprequest.environ)
    def _get_display_name(self, record, field_name, subfield_name='display_name'):
        x = getattr(getattr(record, field_name), subfield_name) if field_name.endswith("_id") else str(getattr(record, field_name))
        return x
    @http.route('/stock_picking/export/xlsx/test', type='http', auth='user')
    def export_xlsx_handler(self, ids):
        output = io.BytesIO()
        ids = [int(id) for id in ids.strip('][').split(', ')]
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet(_('Stock Picking'))
        header_bold = workbook.add_format({'bold': True, 'pattern': 1, 'bg_color': '#AAAAAA'})
        header_plain = workbook.add_format({'pattern': 1, 'bg_color': '#AAAAAA'})
        bold = workbook.add_format({'bold': True})
        filtered_keys = filter(lambda f: f in ('partner_id', 'origin', 'sale_id', 'move_lines.name','move_lines.product_qty','scheduled_date'),request.env['stock.picking']._fields.keys())


        headers = list(filtered_keys)
        for c, header in enumerate(headers):
            worksheet.write(0, c, header, header_bold)
        records = request.env['stock.picking'].browse(ids)
        l = 1
        for record in records:
            for c, field_name in enumerate(headers):
                worksheet.write(l, c, self._get_display_name(record, field_name), header_plain)
            l += 1
        workbook.close()
        xlsx_data = output.getvalue()
        filename = osutil.clean_filename(_("Raport %(title)s (%(model_name)s)", title=_('Problèmes de planning'), model_name="stock_picking"))
        response = request.make_response(xlsx_data,
                                         headers=[('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                                                  ('Content-Disposition', content_disposition(filename + '.xlsx'))],
                                         )

        return response


    @http.route('/stock_picking/export/xlsx', type='http', auth='user')
    def export_xlsx_handler(self, ids):
        ids = [int(id) for id in ids.strip('][').split(', ')]
        records = request.env['stock.picking'].browse(ids)

        workbook = Workbook()

        sheet = workbook.active

        headers = ['Nom du partenaire', 'Document d origine', 'Date de la commande','Description du mouvement de stock','Quantité', 'Date prévue']
        sheet.append(headers)
        header_row = sheet[1]
        for cell in header_row:
            cell.font = Font(bold=True)
        for record in records:
            for st_move in record.move_lines:
                st_name = st_move.name
                st_product_qty = st_move.product_qty
            data = [record.partner_id.name, record.origin, record.date.strftime("%d-%m-%Y"), st_name, st_product_qty, record.scheduled_date.strftime("%d-%m-%Y")]
            sheet.append(data)
        file_stream = io.BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        file_content = file_stream.read()


        filename = 'probleme_planning_data.xlsx'

        # Renvoyer le fichier Excel en tant que réponse HTTP pour le téléchargement automatique
        response = Response(
            file_content,
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', 'attachment; filename=%s' % filename),
            ]
        )
        return response

    #@http.route('/smt_stock/smt_stock', auth='public')
    # def index(self, **kw):
    #     return "Hello, world"
    #
    # @http.route('/smt_stock/smt_stock/objects', auth='public')
    # def list(self, **kw):
    #     return http.request.render('smt_stock.listing', {
    #         'root': '/smt_stock/smt_stock',
    #         'objects': http.request.env['smt_stock.smt_stock'].search([]),
    #     })
    #
    # @http.route('/smt_stock/smt_stock/objects/<model("smt_stock.smt_stock"):obj>', auth='public')
    # def object(self, obj, **kw):
    #     return http.request.render('smt_stock.object', {
    #         'object': obj
    #     })

    # @http.route('/stock_picking/export/xlsx', type='http', auth='user')
    # def export_xlsx_handler(self, ids):
    #     ids = [int(id) for id in ids.strip('][').split(', ')]
    #     records = request.env['stock.picking'].browse(ids)
    #
    #     workbook = Workbook()
    #
    #     sheet = workbook.active
    #
    #     headers = ['Nom du partenaire', 'Document d origine', 'Date de la commande','Description du mouvement de stock','Quantité', 'Date prévue']
    #     sheet.append(headers)
    #     header_row = sheet[1]
    #     for cell in header_row:
    #         cell.font = Font(bold=True)
    #     for record in records:
    #         for st_move in record.move_lines:
    #             st_name = st_move.name
    #             st_product_qty = st_move.product_qty
    #         data = [record.partner_id.name, record.origin, record.date.strftime("%d-%m-%Y"), st_name, st_product_qty, record.scheduled_date.strftime("%d-%m-%Y")]
    #         sheet.append(data)
    #     file_stream = io.BytesIO()
    #     workbook.save(file_stream)
    #     file_stream.seek(0)
    #
    #     file_content = file_stream.read()
    #
    #
    #     filename = 'probleme_planning_data.xlsx'
    #
    #     # Renvoyer le fichier Excel en tant que réponse HTTP pour le téléchargement automatique
    #     response = Response(
    #         file_content,
    #         headers=[
    #             ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
    #             ('Content-Disposition', 'attachment; filename=%s' % filename),
    #         ]
    #     )
    #     return response


